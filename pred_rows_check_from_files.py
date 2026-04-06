#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

FILL_GREEN = PatternFill(fill_type="solid", fgColor="C6EFCE")
FILL_RED = PatternFill(fill_type="solid", fgColor="FFC7CE")
FILL_HEADER = PatternFill(fill_type="solid", fgColor="4F81BD")


def hms_to_seconds(text: str) -> float:
    text = str(text).strip()
    parts = [int(x) for x in text.split(":")]

    if len(parts) == 3:
        h, m, s = parts
        return h * 3600 + m * 60 + s
    elif len(parts) == 2:
        m, s = parts
        return m * 60 + s
    else:
        raise ValueError(f"Unsupported time format: {text}")


def parse_gt_range(text: str):
    parts = re.split(r"\s*-\s*", str(text).strip())
    start_text, end_text = parts
    return hms_to_seconds(start_text), hms_to_seconds(end_text)


def clean_stage_value(value):
    if pd.isna(value):
        return "N/A"

    text = str(value).strip()

    if text == "" or text.lower() == "nan":
        return "N/A"

    # Fix float-like values (e.g., "2.0")
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]

    # Fix Excel date-like strings
    m = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2}) 00:00:00", text)
    if m:
        month = int(m.group(2))
        day = int(m.group(3))
        return f"{month}-{day}"

    return text[:-2]

    m = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2}) 00:00:00", text)
    if m:
        month = int(m.group(2))
        day = int(m.group(3))
        return f"{month}-{day}"

    return text


def overlap_seconds(a_start, a_end, b_start, b_end) -> float:
    return max(0.0, min(a_end, b_end) - max(a_start, b_start))


def style_excel(path: str, sheet_name: str = "pred_rows_checked"):
    wb = load_workbook(path)
    ws = wb[sheet_name]

    headers = [cell.value for cell in ws[1]]
    color_col_idx = headers.index("Color") + 1

    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row in range(2, ws.max_row + 1):
        color_val = ws.cell(row=row, column=color_col_idx).value
        if color_val == "green":
            fill = FILL_GREEN
        elif color_val == "red":
            fill = FILL_RED
        else:
            fill = None

        if fill:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill

    ws.column_dimensions[ws.cell(row=1, column=color_col_idx).column_letter].hidden = True
    wb.save(path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--gt_xlsx", required=True)
    parser.add_argument("--gt_sheet", required=True)
    parser.add_argument("--pred_csv", required=True)
    parser.add_argument("--out_xlsx", required=True)
    args = parser.parse_args()

    gt_df = pd.read_excel(args.gt_xlsx, sheet_name=args.gt_sheet)

    gt_intervals = []
    for _, row in gt_df.iterrows():
        if pd.isna(row["Time"]):
            continue

        gt_start, gt_end = parse_gt_range(row["Time"])
        gt_stage = clean_stage_value(row["Stage"])

        gt_intervals.append((gt_start, gt_end, gt_stage, row["Time"]))

    pred_df = pd.read_csv(args.pred_csv)
    pred_df = pred_df[pred_df["pred_label"] == 1]

    results = []

    for _, row in pred_df.iterrows():
        pred_start = float(row["start_sec_approx"])
        pred_end = float(row["end_sec_approx"])
        pred_len = pred_end - pred_start

        max_overlap = 0
        stages = []
        gt_times = []

        for gt_start, gt_end, gt_stage, gt_time in gt_intervals:
            ov = overlap_seconds(pred_start, pred_end, gt_start, gt_end)

            if ov > max_overlap:
                max_overlap = ov

            if ov > 0:
                if gt_stage:
                    stages.append(gt_stage)
                gt_times.append(gt_time)

        overlap_ratio = max_overlap / pred_len if pred_len > 0 else 0
        detected = max_overlap > 0

        results.append({
            "window_index": row["window_index"],
            "prob": row["prob"],
            "start_hms_approx": row["start_hms_approx"],
            "end_hms_approx": row["end_hms_approx"],
            "Overlap_ratio": round(overlap_ratio, 3),
            "Stage": "; ".join(set(stages)),
            "GT_Time": "; ".join(set(gt_times)),
            "Color": "green" if detected else "red"
        })

    df = pd.DataFrame(results)

    with pd.ExcelWriter(args.out_xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="pred_rows_checked", index=False)

    style_excel(args.out_xlsx)
    print("[OK] Saved:", args.out_xlsx)


if __name__ == "__main__":
    main()
